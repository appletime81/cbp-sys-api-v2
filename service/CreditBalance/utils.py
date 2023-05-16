from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import pandas as pd
from pprint import pprint
import numpy as np
from fastapi.responses import FileResponse


def generate_credit_balance_report(df: pd.DataFrame, partyName: str = "KT"):
    org_col_names = df.columns.tolist()
    converted_col_names = [
        "海纜名稱",
        "海纜作業",
        "記帳段號",
        "Invoice No.",
        "Issue Date",
        "CN no.",
        "Issue Date",
        "Description",
        "Debit",
        "Credit",
        "Balance",
    ]
    col_names_map = dict(zip(org_col_names, converted_col_names))

    # start to generate excel from openpyxl
    max_col = len(converted_col_names)
    max_row = len(df.index) + 2  # include header and comment(Currency:USD, col names)

    wb = Workbook()
    ws = wb.active
    ws.title = "Credit Balance Report"

    # let all grid are transparent
    ws.sheet_view.showGridLines = False

    # set column width
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 50
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 20

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if row == 1:
                # no borderline(transparent)
                ws.cell(row=row, column=col).border = Border(
                    left=Side(border_style="thin", color="00FFFFFF"),
                    right=Side(border_style="thin", color="00FFFFFF"),
                    top=Side(border_style="thin", color="00FFFFFF"),
                    bottom=Side(border_style="thin", color="00FFFFFF"),
                )
                if col == 1:
                    ws.cell(
                        row=row, column=col
                    ).value = (
                        f"{df.iloc[0, 0]} {df.iloc[0, 1]}_Credit Balance_{partyName}"
                    )
                    # set font be Bold
                    ws.cell(row=row, column=col).font = Font(bold=True)
                if col == 11:
                    # set comment
                    ws.cell(row=row, column=col).value = "Currency: USD"
            elif row == 2:
                # set header
                ws.cell(row=row, column=col).value = converted_col_names[col - 1]
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFC000", end_color="FFC000", fill_type="solid"
                )
            elif row > 2:
                # set data
                ws.cell(row=row, column=col).value = df.iloc[row - 3, col - 1]
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

                # set borderline and color is black
                ws.cell(row=row, column=col).border = Border(
                    left=Side(border_style="thin", color="FF000000"),
                    right=Side(border_style="thin", color="FF000000"),
                    top=Side(border_style="thin", color="FF000000"),
                    bottom=Side(border_style="thin", color="FF000000"),
                )

                if col > 9:
                    # set align right
                    ws.cell(row=row, column=col).alignment = Alignment(
                        horizontal="right", vertical="center"
                    )

    # set row = max_row + 1, col = A to J merged and set value "Sub-total Balance", and K = sum of Balance
    ws.merge_cells(
        start_row=max_row + 1, start_column=1, end_row=max_row + 1, end_column=10
    )
    ws.cell(row=max_row + 1, column=1).value = "Sub-total Balance"
    ws.cell(row=max_row + 1, column=11).value = df.iloc[max_row - 3, max_col - 1]

    # set all number value have ,
    for row in range(1, max_row + 2):
        for col in range(1, max_col + 1):
            if ws.cell(row=row, column=col).value is not None and isinstance(
                ws.cell(row=row, column=col).value, (float)
            ):
                ws.cell(row=row, column=col).number_format = "#,##0.00"
            # 如果沒有小數點 也要有三位一個逗號
            elif ws.cell(row=row, column=col).value is not None and isinstance(
                ws.cell(row=row, column=col).value, (np.int64)
            ):
                ws.cell(row=row, column=col).number_format = "#,##0.00"

    # set all fontstyle are Arial
    for row in range(1, max_row + 2):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).font = Font(name="Arial")
            if row > max_row:
                # set font Color blue and bold
                ws.cell(row=row, column=col).font = Font(
                    color="FF0000FF", bold=True, name="Arial"
                )

                # set "ws.cell(row=max_row + 1, column=1).font = Font(color="FF0000FF", bold=True)" align to left
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="right", vertical="center"
                )

    file_name = f"CB歷程.xlsx"
    wb.save(file_name)
    resp = FileResponse(path=f"{file_name}", filename=f"{file_name}")
    return resp
